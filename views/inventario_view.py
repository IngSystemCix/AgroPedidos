from PIL import Image
import customtkinter as ctk
from tkinter import ttk, filedialog
import os
import pandas as pd
from datetime import datetime
from services.product_service import get_all_products
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

class InventarioView(ctk.CTkFrame):
    def __init__(self, master, usuario, navigate):
        super().__init__(master)
        self.master = master
        self.usuario = usuario
        self.navigate = navigate
        self.configure(fg_color="#f3fdf2")
        self.pack(fill="both", expand=True)
        self.create_widgets()

    def create_widgets(self):
        # -------- HEADER --------
        header = ctk.CTkFrame(self, fg_color="#ffffff", height=80)
        header.pack(fill="x", side="top")

        try:
            logo_img = Image.open("./resources/images/logo.png")
            logo = ctk.CTkImage(light_image=logo_img, dark_image=logo_img, size=(60, 60))
            ctk.CTkLabel(header, image=logo, text="").pack(side="left", padx=10, pady=10)
        except:
            pass

        ctk.CTkLabel(header, text="AGROPEDIDOS", font=("Segoe UI", 28, "bold"), text_color="#1a8341").pack(side="left", padx=10)

        user_section = ctk.CTkFrame(header, fg_color="transparent")
        user_section.pack(side="right", padx=20)
        ctk.CTkLabel(user_section, text="👨‍💼", font=("Segoe UI", 18)).pack(side="left")
        ctk.CTkLabel(user_section, text="Administrador", font=("Segoe UI", 14)).pack(side="left", padx=5)
        ctk.CTkLabel(user_section, text=self.usuario.username, font=("Segoe UI", 14, "bold")).pack(side="left", padx=5)
        ctk.CTkButton(
            user_section,
            text="Cerrar sesión",
            width=100,
            fg_color="#ff4d4d",
            hover_color="#cc0000",
            command=self.navigate_logout
        ).pack(side="left", padx=10)

        # -------- NAVEGACIÓN --------
        nav_frame = ctk.CTkFrame(self, fg_color="#f3fdf2")
        nav_frame.pack(pady=(5, 0))
        secciones = [("Catálogo", "catalogo"), ("Gestión de Productos", "gestion"), ("Inventario", "inventario"), ("Ventas", "ventas")]
        for label, destino in secciones:
            ctk.CTkButton(nav_frame, text=label, width=150, command=lambda d=destino: self.navigate(d)).pack(side="left", padx=10)

        # -------- CONTENIDO --------
        top_frame = ctk.CTkFrame(self, fg_color="#f3fdf2")
        top_frame.pack(fill="x", padx=20, pady=(10, 0))

        self.search_entry = ctk.CTkEntry(top_frame, placeholder_text="Buscar producto...", width=300)
        self.search_entry.pack(side="right", padx=(0, 10))
        self.search_entry.bind("<KeyRelease>", lambda event: self.filtrar_tabla())

        export_btn = ctk.CTkButton(top_frame, text="Exportar", width=120, command=self.exportar_excel)
        export_btn.pack(side="right", padx=5)

        # Tabla
        tabla_frame = ctk.CTkFrame(self, fg_color="white")
        tabla_frame.pack(fill="both", expand=True, padx=20, pady=10)

        self.tree = ttk.Treeview(tabla_frame, columns=("Producto", "Stock", "Precio", "Valor", "Estado"), show="headings")
        for col in self.tree["columns"]:
            self.tree.heading(col, text=col)
            self.tree.column(col, anchor="center")
        self.tree.pack(fill="both", expand=True, pady=10, padx=10)

        self.cargar_datos()

    def cargar_datos(self):
        self.productos = get_all_products()
        self.mostrar_productos(self.productos)

    def mostrar_productos(self, productos):
        for item in self.tree.get_children():
            self.tree.delete(item)

        for p in productos:
            valor_total = p.price * p.stock
            if p.stock == 0:
                estado = "Agotado"
            elif p.stock < 50:
                estado = "Bajo"
            else:
                estado = "Disponible"
            tag = estado.lower()
            self.tree.insert("", "end", values=(p.name, p.stock, f"S/ {p.price:.2f}", f"S/ {valor_total:.2f}", estado), tags=(tag,))

        self.tree.tag_configure("agotado", background="#fdd")
        self.tree.tag_configure("bajo", background="#fff2cc")
        self.tree.tag_configure("disponible", background="#d9ead3")

    def filtrar_tabla(self):
        query = self.search_entry.get().lower()
        filtrados = [p for p in self.productos if query in p.name.lower()]
        self.mostrar_productos(filtrados)

    def exportar_excel(self):
        data = []
        for p in self.productos:
            valor_total = p.price * p.stock
            if p.stock == 0:
                estado = "Agotado"
            elif p.stock < 50:
                estado = "Bajo"
            else:
                estado = "Disponible"
            data.append([p.name, p.stock, p.price, valor_total, estado])

        df = pd.DataFrame(data, columns=["Producto", "Stock Actual", "Precio Unitario", "Valor Total", "Estado"])
        fecha_actual = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"inventario_exportado_{fecha_actual}.xlsx"
        save_path = os.path.join("resources", filename)
        df.to_excel(save_path, index=False)

        # Estilos en Excel
        wb = load_workbook(save_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                estado = cell.value
                if estado == "Agotado":
                    cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                elif estado == "Bajo":
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                elif estado == "Disponible":
                    cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        wb.save(save_path)

        # Modal de éxito
        confirm_popup = ctk.CTkToplevel(self)
        confirm_popup.title("Éxito")
        confirm_popup.geometry("300x150")
        confirm_popup.grab_set()

        ctk.CTkLabel(confirm_popup, text="✅ Datos exportados correctamente", font=("Segoe UI", 14)).pack(pady=30)
        ctk.CTkButton(confirm_popup, text="OK", width=80, command=confirm_popup.destroy).pack()

    def navigate_logout(self):
        self.navigate("logout")
